"""
Phase V2-1.4 — Simulated Data Generation (V2 rewrite)
======================================================
Generates:
  - data/seed/master_data.csv          10 diverse customer records (customer_id PK)
  - data/seed/dlu_metadata.csv        one row per simulated file (MD5 + file_path only)
  - data/simulated_files/             ~25 breach files (txt, xlsx, csv, xls)
  - data/TEXT/{md5[:3]}/{md5}.ext     indexed copy of every breach file

V2 changes from V1:
  - master_data.csv replaces master_pii.csv
  - customer_id (INT 1-10) replaces ID string (C001-C010) as primary key
  - dlu_metadata.csv has only MD5 and file_path (no GUID, caseName, etc.)
  - file_path uses forward-slash: data/TEXT/{md5[:3]}/{md5}.ext

Run from the project root:
    python scripts/generate_simulated_data.py

The script is idempotent — it wipes and recreates the output dirs on every run.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import os
import shutil
import struct
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

try:
    import xlwt
    _HAS_XLWT = True
except ImportError:
    _HAS_XLWT = False

# ---------------------------------------------------------------------------
# Paths (relative to project root — script is run from there)
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
SEED_DIR = DATA_DIR / "seed"
SIM_FILES_DIR = DATA_DIR / "simulated_files"
TEXT_DIR = DATA_DIR / "TEXT"
MASTER_DATA_CSV = SEED_DIR / "master_data.csv"
DLU_METADATA_CSV = SEED_DIR / "dlu_metadata.csv"

# ---------------------------------------------------------------------------
# Customer master data
# V2: customer_id is INT (1..10), all other fields identical to V1
# ---------------------------------------------------------------------------

# fmt: off
CUSTOMERS: list[dict[str, Any]] = [
    {
        "customer_id": 1,
        "Fullname": "Robert O'Brien",
        "FirstName": "Robert",
        "LastName": "O'Brien",
        "DOB": "1975-03-22",
        "SSN": "523-45-7891",
        "DriversLicense": "MA-O8834521",
        "Address1": "14 Clover Lane",
        "Address2": "Apt 3B",
        "Address3": "",
        "ZipCode": "02101",
        "City": "Boston",
        "State": "MA",
        "Country": "USA",
    },
    {
        "customer_id": 2,
        "Fullname": "Jennifer Smith-Jones",
        "FirstName": "Jennifer",
        "LastName": "Smith-Jones",
        "DOB": "1988-11-04",
        "SSN": "412-67-3309",
        "DriversLicense": "TX-J9921034",
        "Address1": "8820 Pecan Street",
        "Address2": "",
        "Address3": "",
        "ZipCode": "78701",
        "City": "Austin",
        "State": "TX",
        "Country": "USA",
    },
    {
        "customer_id": 3,
        "Fullname": "Karthik Chekuri",
        "FirstName": "Karthik",
        "LastName": "Chekuri",
        "DOB": "1992-07-15",
        "SSN": "634-21-8805",
        "DriversLicense": "CA-K3344112",
        "Address1": "2200 Wilshire Blvd",
        "Address2": "Suite 510",
        "Address3": "Floor 5",
        "ZipCode": "90057",
        "City": "Los Angeles",
        "State": "CA",
        "Country": "USA",
    },
    {
        "customer_id": 4,
        "Fullname": "Maria Rodriguez",
        "FirstName": "Maria",
        "LastName": "Rodriguez",
        "DOB": "1980-06-30",
        "SSN": "291-88-4451",
        "DriversLicense": "FL-M2288991",
        "Address1": "305 Biscayne Way",
        "Address2": "",
        "Address3": "",
        "ZipCode": "33101",
        "City": "Miami",
        "State": "FL",
        "Country": "",
    },
    {
        "customer_id": 5,
        "Fullname": "Ahmed Hassan",
        "FirstName": "Ahmed",
        "LastName": "Hassan",
        "DOB": "1970-01-19",
        "SSN": "785-33-6624",
        "DriversLicense": "NY-A1198834",
        "Address1": "77 Atlantic Ave",
        "Address2": "Unit 12",
        "Address3": "",
        "ZipCode": "11201",
        "City": "Brooklyn",
        "State": "NY",
        "Country": "USA",
    },
    {
        "customer_id": 6,
        "Fullname": "Priya Patel",
        "FirstName": "Priya",
        "LastName": "Patel",
        "DOB": "1995-09-08",
        "SSN": "348-56-7712",
        "DriversLicense": "IL-P7712349",
        "Address1": "456 Michigan Avenue",
        "Address2": "",
        "Address3": "",
        "ZipCode": "60611",
        "City": "Chicago",
        "State": "IL",
        "Country": "USA",
    },
    {
        "customer_id": 7,
        "Fullname": "James Kim",
        "FirstName": "James",
        "LastName": "Kim",
        "DOB": "1983-04-27",
        "SSN": "901-24-5567",
        "DriversLicense": "WA-J5521190",
        "Address1": "3912 Pine Street",
        "Address2": "Apt 201",
        "Address3": "",
        "ZipCode": "98101",
        "City": "Seattle",
        "State": "WA",
        "Country": "USA",
    },
    {
        "customer_id": 8,
        "Fullname": "Linda Thornberry",
        "FirstName": "Linda",
        "LastName": "Thornberry",
        "DOB": "1967-12-11",
        "SSN": "677-09-3381",
        "DriversLicense": "OH-L3381677",
        "Address1": "88 Maple Drive",
        "Address2": "",
        "Address3": "",
        "ZipCode": "44101",
        "City": "Cleveland",
        "State": "OH",
        "Country": "",
    },
    {
        "customer_id": 9,
        "Fullname": "Carlos Reyes-Morales",
        "FirstName": "Carlos",
        "LastName": "Reyes-Morales",
        "DOB": "1990-02-14",
        "SSN": "556-77-2290",
        "DriversLicense": "AZ-C2290556",
        "Address1": "1101 Saguaro Blvd",
        "Address2": "Unit B",
        "Address3": "",
        "ZipCode": "85001",
        "City": "Phoenix",
        "State": "AZ",
        "Country": "USA",
    },
    {
        "customer_id": 10,
        "Fullname": "Susan Whitfield",
        "FirstName": "Susan",
        "LastName": "Whitfield",
        "DOB": "1978-08-05",
        "SSN": "143-62-9908",
        "DriversLicense": "GA-S9908143",
        "Address1": "2700 Peachtree Road NE",
        "Address2": "Suite 100",
        "Address3": "Building C",
        "ZipCode": "30305",
        "City": "Atlanta",
        "State": "GA",
        "Country": "USA",
    },
]
# fmt: on

# Convenience lookup by integer customer_id
CUST: dict[int, dict[str, Any]] = {c["customer_id"]: c for c in CUSTOMERS}


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def md5_of_bytes(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def ssn_nodash(ssn: str) -> str:
    """Return SSN without dashes: '523-45-7891' -> '523457891'."""
    return ssn.replace("-", "")


def dob_us(dob_iso: str) -> str:
    """Convert ISO date to US format: '1975-03-22' -> '03/22/1975'."""
    y, m, d = dob_iso.split("-")
    return f"{m}/{d}/{y}"


def dob_eu(dob_iso: str) -> str:
    """Convert ISO date to European format: '1975-03-22' -> '22/03/1975'."""
    y, m, d = dob_iso.split("-")
    return f"{d}/{m}/{y}"


def last_first(customer: dict) -> str:
    """Return 'LastName, FirstName' ordering."""
    return f"{customer['LastName']}, {customer['FirstName']}"


# ---------------------------------------------------------------------------
# Text-file generators
# ---------------------------------------------------------------------------

def make_appointment_note(c1: dict, c2: dict) -> str:
    """Appointment notes document — two patients."""
    return f"""PATIENT APPOINTMENT NOTES
Facility: Riverside Medical Center
Date: {dob_us(c1["DOB"])[:5].replace("/", "-")}2024

---
Patient: {c1["Fullname"]}
Date of Birth: {dob_us(c1["DOB"])}
SSN: {c1["SSN"]}
Address: {c1["Address1"]}, {c1["City"]}, {c1["State"]} {c1["ZipCode"]}
Driver's License: {c1["DriversLicense"]}
Reason for Visit: Annual wellness checkup
Notes: Patient reports mild fatigue. Labs ordered. Follow-up in 6 weeks.

---
Patient: {c2["Fullname"]}
Date of Birth: {c2["DOB"]}
SSN: {ssn_nodash(c2["SSN"])}
Address: {c2["Address1"]}, {c2["City"]}, {c2["State"]} {c2["ZipCode"]}
Driver's License: {c2["DriversLicense"]}
Reason for Visit: Prescription renewal
Notes: Renewed Lisinopril 10mg. No change in dosage. BP within normal range.
"""


def make_hr_onboarding(c1: dict) -> str:
    """HR onboarding form — single employee."""
    return f"""HUMAN RESOURCES ONBOARDING FORM
Company: Meridian Technologies LLC
Form Completed: 2024-01-15

EMPLOYEE INFORMATION
Full Name: {c1["Fullname"]}
First Name: {c1["FirstName"]}
Last Name: {c1["LastName"]}
Date of Birth: {dob_us(c1["DOB"])}
Social Security Number: {c1["SSN"]}
Driver License No.: {c1["DriversLicense"]}
Home Address: {c1["Address1"]}{" " + c1["Address2"] if c1["Address2"] else ""}
City: {c1["City"]}
State: {c1["State"]}
ZIP: {c1["ZipCode"]}

EMPLOYMENT DETAILS
Position: Senior Analyst
Department: Operations
Start Date: 02/01/2024
Salary Grade: G7

EMERGENCY CONTACT
Name: Jane Doe
Relationship: Spouse
Phone: 555-000-1234

I certify that the above information is accurate.
Signature: ___________________  Date: ______________
"""


def make_insurance_claim(c1: dict, c2: dict) -> str:
    """Insurance claim document with one intentional misspelling."""
    # Introduce misspelling for Rodriguez / use last-first for second customer
    name1_variant = c1["Fullname"].replace("Rodriguez", "Rodgriguez")
    name2_variant = last_first(c2) if c2["LastName"] else c2["Fullname"]
    return f"""INSURANCE CLAIM FORM
Policy Provider: BlueStar Insurance Group
Claim Date: 2024-03-10
Claim Number: CLM-2024-88741

CLAIMANT 1
Name: {name1_variant}
SSN: {ssn_nodash(c1["SSN"])}
Date of Birth: {dob_eu(c1["DOB"])}
Address: {c1["Address1"]}, {c1["City"]}, {c1["State"]} {c1["ZipCode"]}
Claim Type: Medical — Inpatient
Amount Claimed: $4,250.00

CLAIMANT 2
Name: {name2_variant}
SSN: {c2["SSN"]}
Date of Birth: {dob_us(c2["DOB"])}
Address: {c2["Address1"]}, {c2["City"]}, {c2["State"]} {c2["ZipCode"]}
Claim Type: Dental — Orthodontics
Amount Claimed: $1,800.00

Adjuster Notes: Both claims under review pending documentation.
"""


def make_tax_w2(c1: dict) -> str:
    """Simplified W-2 tax form."""
    return f"""W-2 WAGE AND TAX STATEMENT — Tax Year 2023
Employer: Apex Global Solutions Inc.
EIN: 47-1234567
Employer Address: 500 Corporate Plaza, Denver, CO 80202

EMPLOYEE INFORMATION
Employee Name: {c1["Fullname"]}
SSN: {c1["SSN"]}
Address: {c1["Address1"]}{", " + c1["Address2"] if c1["Address2"] else ""}
         {c1["City"]}, {c1["State"]} {c1["ZipCode"]}

BOX 1 — Wages, Tips: $82,500.00
BOX 2 — Federal Tax Withheld: $16,320.00
BOX 3 — Social Security Wages: $82,500.00
BOX 4 — Social Security Tax: $5,115.00
BOX 12 — Code D (401k): $6,500.00
BOX 16 — State Wages: $82,500.00
BOX 17 — State Tax: $4,537.50
State: {c1["State"]}

This document is provided for tax purposes only.
"""


def make_benefits_enrollment(c1: dict, c2: dict, c3: dict) -> str:
    """Benefits enrollment form with three employees and date variation."""
    return f"""ANNUAL BENEFITS ENROLLMENT FORM
Open Enrollment Period: November 1 – November 30, 2023
HR System: PeopleFirst HRIS

PARTICIPANT 1
Employee ID: {c1["customer_id"]}
Name: {c1["Fullname"]}
DOB: {c1["DOB"]}
SSN: {c1["SSN"]}
Coverage Selected: Medical PPO + Dental + Vision
Effective Date: 01/01/2024

PARTICIPANT 2
Employee ID: {c2["customer_id"]}
Name: {last_first(c2)}
DOB: {dob_us(c2["DOB"])}
SSN: {ssn_nodash(c2["SSN"])}
Coverage Selected: HMO + Dental
Effective Date: 01/01/2024

PARTICIPANT 3
Employee ID: {c3["customer_id"]}
Name: {c3["Fullname"]}
DOB: {dob_eu(c3["DOB"])}
SSN: {c3["SSN"]}
Coverage Selected: Medical PPO
Effective Date: 01/01/2024

Authorized HR Signature: _______________________
"""


def make_client_intake(c1: dict) -> str:
    """Law-firm client intake form."""
    return f"""CLIENT INTAKE FORM
Firm: Harrington & Sloane, LLP
Date: {dob_us(c1["DOB"])[:2]}/2024

CLIENT DETAILS
Full Name: {c1["Fullname"]}
Date of Birth: {c1["DOB"]}
Social Security: {c1["SSN"]}
Driver License: {c1["DriversLicense"]}
Primary Address: {c1["Address1"]}
Secondary Address: {c1["Address2"] if c1["Address2"] else "N/A"}
City / State / ZIP: {c1["City"]}, {c1["State"]} {c1["ZipCode"]}
Country: {c1["Country"] if c1["Country"] else "United States"}

MATTER TYPE: Estate Planning
Assigned Attorney: M. Harrington, Esq.
Conflict Check: CLEARED
Fee Agreement Signed: YES

Intake Coordinator: ______________  Date: ___________
"""


def make_payroll_register(customers: list[dict]) -> str:
    """Payroll register for a list of employees."""
    lines = [
        "PAYROLL REGISTER",
        "Pay Period: 2024-Q1 (January–March)",
        "Entity: Orion Manufacturing Corp.",
        f"{'Emp ID':<8} {'Name':<25} {'SSN':<14} {'Gross Pay':>10} {'Net Pay':>10} {'State':<4}",
        "-" * 80,
    ]
    for c in customers:
        gross = 21_000 + (hash(c["SSN"]) % 5_000)
        net = int(gross * 0.72)
        # Use abbreviation for some, full name for others
        name_col = f"J. {c['LastName']}" if c["FirstName"].startswith("J") else c["Fullname"]
        # Alternate SSN format: odd customer_id = dashed, even = nodash
        ssn_col = ssn_nodash(c["SSN"]) if c["customer_id"] % 2 == 0 else c["SSN"]
        lines.append(
            f"{c['customer_id']:<8} {name_col:<25} {ssn_col:<14} ${gross:>9,.2f} ${net:>9,.2f} {c['State']:<4}"
        )
    lines += [
        "-" * 80,
        "Payroll Administrator: ____________________",
        "Approval Date: ____________________________",
    ]
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# CSV-file generators
# ---------------------------------------------------------------------------

def make_csv_employee_directory(customers: list[dict]) -> list[list[str]]:
    """Simple CSV employee directory."""
    headers = ["EmployeeID", "FullName", "SSN", "DOB", "Address", "City", "State", "ZipCode"]
    rows = [headers]
    for c in customers:
        rows.append([
            str(c["customer_id"]),
            c["Fullname"],
            c["SSN"],
            dob_us(c["DOB"]),
            c["Address1"],
            c["City"],
            c["State"],
            c["ZipCode"],
        ])
    return rows


def make_csv_payroll_export(customers: list[dict]) -> list[list[str]]:
    """CSV payroll export with mixed SSN formats."""
    headers = ["ID", "LastFirst", "SSN", "DOB_ISO", "Gross", "Deductions", "Net", "State"]
    rows = [headers]
    for i, c in enumerate(customers):
        gross = 7_500 + (hash(str(c["customer_id"])) % 2_000)
        deductions = int(gross * 0.28)
        net = gross - deductions
        ssn_col = c["SSN"] if i % 2 == 0 else ssn_nodash(c["SSN"])
        rows.append([
            str(c["customer_id"]),
            last_first(c),
            ssn_col,
            c["DOB"],
            f"{gross:.2f}",
            f"{deductions:.2f}",
            f"{net:.2f}",
            c["State"],
        ])
    return rows


def make_csv_insurance_roster(customers: list[dict]) -> list[list[str]]:
    """CSV insurance roster with intentional misspelling."""
    headers = ["PolicyHolder", "DOB", "SSN", "DriversLicense", "Address", "ZipCode", "State"]
    rows = [headers]
    for c in customers:
        # Introduce misspelling for Rodriguez
        name = c["Fullname"].replace("Rodriguez", "Rodgriguez")
        rows.append([
            name,
            dob_eu(c["DOB"]),
            ssn_nodash(c["SSN"]),
            c["DriversLicense"],
            c["Address1"],
            c["ZipCode"],
            c["State"],
        ])
    return rows


def make_csv_client_list(customers: list[dict]) -> list[list[str]]:
    """CSV client list with last-first name ordering."""
    headers = ["ClientID", "Name_LastFirst", "SSN", "DOB_US", "City", "State"]
    rows = [headers]
    for c in customers:
        rows.append([
            str(c["customer_id"]),
            last_first(c),
            c["SSN"],
            dob_us(c["DOB"]),
            c["City"],
            c["State"],
        ])
    return rows


# ---------------------------------------------------------------------------
# XLSX generators
# ---------------------------------------------------------------------------

def make_xlsx_hr_form(c1: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HR Form"
    ws.append(["Field", "Value"])
    data = [
        ("Employee Name", c1["Fullname"]),
        ("Date of Birth", dob_us(c1["DOB"])),
        ("SSN", c1["SSN"]),
        ("Driver License", c1["DriversLicense"]),
        ("Address Line 1", c1["Address1"]),
        ("Address Line 2", c1["Address2"]),
        ("City", c1["City"]),
        ("State", c1["State"]),
        ("ZIP Code", c1["ZipCode"]),
        ("Country", c1["Country"]),
        ("Department", "Finance"),
        ("Manager", "R. Davis"),
        ("Start Date", "2024-02-01"),
    ]
    for row in data:
        ws.append(list(row))
    return wb


def make_xlsx_payroll(customers: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Q1"
    ws.append(["EmpID", "Name", "SSN", "DOB", "GrossPay", "NetPay", "State"])
    for c in customers:
        gross = 21_000 + (hash(c["SSN"]) % 5_000)
        net = int(gross * 0.72)
        ws.append([
            c["customer_id"],
            c["Fullname"],
            ssn_nodash(c["SSN"]),
            c["DOB"],
            gross,
            net,
            c["State"],
        ])
    return wb


def make_xlsx_benefits(customers: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benefits Enrollment"
    ws.append(["ID", "LastFirst", "DOB_US", "SSN", "Plan", "Effective"])
    plans = ["Medical PPO", "HMO", "Medical PPO + Dental", "Dental Only", "Vision"]
    for i, c in enumerate(customers):
        ws.append([
            c["customer_id"],
            last_first(c),
            dob_us(c["DOB"]),
            c["SSN"],
            plans[i % len(plans)],
            "01/01/2024",
        ])
    return wb


def make_xlsx_client_intake(customers: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Intake"
    ws.append(["ClientID", "FullName", "DOB", "SSN", "DriversLicense",
               "Address1", "Address2", "City", "State", "ZIP", "Country"])
    for c in customers:
        ws.append([
            c["customer_id"], c["Fullname"], c["DOB"], c["SSN"], c["DriversLicense"],
            c["Address1"], c["Address2"], c["City"], c["State"], c["ZipCode"], c["Country"],
        ])
    return wb


# ---------------------------------------------------------------------------
# Minimal BIFF5 XLS writer (pure Python stdlib — no xlwt required)
# ---------------------------------------------------------------------------

def _biff_record(record_type: int, data: bytes) -> bytes:
    """Pack a single BIFF record: type(2) + length(2) + data."""
    return struct.pack("<HH", record_type, len(data)) + data


def _biff_string(s: str) -> bytes:
    """Encode a string as a BIFF5 byte-string (1-byte length + bytes)."""
    encoded = s.encode("latin-1", errors="replace")
    return struct.pack("<B", len(encoded)) + encoded


def _write_biff_workbook(sheet_name: str, rows: list[list]) -> bytes:
    """
    Write a minimal BIFF5 .xls workbook with one sheet.
    Falls back to SYLK format if xlwt is not installed.
    """
    if _HAS_XLWT:
        wb = xlwt.Workbook()
        ws = wb.add_sheet(sheet_name)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Pure-Python fallback: SYLK format that Excel recognises with .xls extension
    lines = ["ID;PWXL;N;E"]
    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            col_letter = chr(ord("A") + c) if c < 26 else f"A{chr(ord('A') + c - 26)}"
            if isinstance(val, (int, float)):
                lines.append(f"C;Y{r + 1};X{c + 1};N{val}")
            else:
                s = str(val).replace(";", ",") if val else ""
                lines.append(f"C;Y{r + 1};X{c + 1};K\"{s}\"")
    lines.append("E")
    return "\n".join(lines).encode("latin-1", errors="replace")


# ---------------------------------------------------------------------------
# XLS data builders — return list-of-lists for our writer
# ---------------------------------------------------------------------------

def make_xls_payroll(customers: list[dict]) -> tuple[str, list[list]]:
    headers = ["EmpID", "Name", "SSN", "DOB_EU", "Gross", "Net", "State"]
    rows = [headers]
    for c in customers:
        gross = 21_000 + (hash(c["SSN"]) % 5_000)
        net = int(gross * 0.72)
        rows.append([
            c["customer_id"], c["Fullname"], c["SSN"],
            dob_eu(c["DOB"]), gross, net, c["State"],
        ])
    return "Payroll", rows


def make_xls_employee_directory(customers: list[dict]) -> tuple[str, list[list]]:
    headers = ["ID", "FirstName", "LastName", "DOB", "SSN", "DriversLicense",
               "Address", "City", "State", "ZIP"]
    rows = [headers]
    for c in customers:
        rows.append([
            c["customer_id"], c["FirstName"], c["LastName"],
            dob_us(c["DOB"]), ssn_nodash(c["SSN"]), c["DriversLicense"],
            c["Address1"], c["City"], c["State"], c["ZipCode"],
        ])
    return "Directory", rows


def make_xls_insurance_claims(customers: list[dict]) -> tuple[str, list[list]]:
    headers = ["ClaimID", "PolicyHolder", "SSN", "DOB", "ClaimType", "Amount", "State"]
    rows = [headers]
    claim_types = ["Medical", "Dental", "Vision", "Pharmacy", "Mental Health"]
    for row_idx, c in enumerate(customers, start=1):
        claim_id = f"CLM-2024-{1000 + row_idx:04d}"
        name = c["Fullname"].replace("Rodriguez", "Rodgriguez")  # intentional misspelling
        amount = 1_500 + (hash(str(c["customer_id"])) % 3_000)
        rows.append([
            claim_id, name, c["SSN"], c["DOB"],
            claim_types[row_idx % len(claim_types)],
            amount, c["State"],
        ])
    return "Claims", rows


# ---------------------------------------------------------------------------
# Write helpers — write to simulated_files, compute MD5, copy to TEXT/
# V2: returns only MD5 + file_path (forward-slash convention)
# ---------------------------------------------------------------------------

def write_bytes_and_index(filename: str, content: bytes) -> dict[str, str]:
    """
    Write `content` to data/simulated_files/{filename},
    compute MD5, copy to data/TEXT/{md5[:3]}/{md5}.ext.

    Returns V2 metadata dict with only MD5 and file_path.
    file_path uses forward-slash: data/TEXT/{md5[:3]}/{md5}.ext
    """
    sim_path = SIM_FILES_DIR / filename
    sim_path.write_bytes(content)

    md5 = md5_of_bytes(content)
    ext = Path(filename).suffix  # e.g. ".txt"
    text_subdir = TEXT_DIR / md5[:3]
    text_subdir.mkdir(parents=True, exist_ok=True)
    text_path = text_subdir / f"{md5}{ext}"
    shutil.copy2(str(sim_path), str(text_path))

    # V2: forward-slash file_path convention
    file_path = f"data/TEXT/{md5[:3]}/{md5}{ext}"
    return {
        "MD5": md5,
        "file_path": file_path,
    }


def write_text_file(filename: str, content: str) -> dict[str, str]:
    return write_bytes_and_index(filename, content.encode("utf-8"))


def write_csv_file(filename: str, rows: list[list[str]]) -> dict[str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return write_bytes_and_index(filename, buf.getvalue().encode("utf-8"))


def write_xlsx_file(filename: str, wb: openpyxl.Workbook) -> dict[str, str]:
    buf = io.BytesIO()
    wb.save(buf)
    return write_bytes_and_index(filename, buf.getvalue())


def write_xls_file(filename: str, sheet_name_and_rows: tuple[str, list[list]]) -> dict[str, str]:
    sheet_name, rows = sheet_name_and_rows
    content = _write_biff_workbook(sheet_name, rows)
    return write_bytes_and_index(filename, content)


# ---------------------------------------------------------------------------
# Main generation logic
# ---------------------------------------------------------------------------

def generate_all() -> None:
    # ------------------------------------------------------------------
    # 1. Clean and recreate output directories
    # ------------------------------------------------------------------
    for d in (SIM_FILES_DIR, TEXT_DIR, SEED_DIR):
        if d.exists():
            shutil.rmtree(str(d))
        d.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # 2. Write master_data.csv (V2: customer_id PK, INT 1-10)
    # ------------------------------------------------------------------
    master_data_columns = [
        "customer_id",
        "Fullname", "FirstName", "LastName", "DOB", "SSN",
        "DriversLicense", "Address1", "Address2", "Address3",
        "ZipCode", "City", "State", "Country",
    ]
    with open(MASTER_DATA_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=master_data_columns)
        writer.writeheader()
        writer.writerows(CUSTOMERS)

    print(f"  master_data.csv written: {len(CUSTOMERS)} rows")

    # ------------------------------------------------------------------
    # 3. Generate ~25 breach files and collect V2 metadata
    # ------------------------------------------------------------------
    metadata_rows: list[dict[str, str]] = []

    c = CUST  # shorthand (integer keys: 1..10)

    # --- .txt files (14) ---
    metadata_rows.append(write_text_file(
        "appointment_notes_mar2024.txt",
        make_appointment_note(c[1], c[5]),
    ))
    metadata_rows.append(write_text_file(
        "appointment_notes_apr2024.txt",
        make_appointment_note(c[6], c[7]),
    ))
    metadata_rows.append(write_text_file(
        "hr_onboarding_obrien.txt",
        make_hr_onboarding(c[1]),
    ))
    metadata_rows.append(write_text_file(
        "hr_onboarding_chekuri.txt",
        make_hr_onboarding(c[3]),
    ))
    metadata_rows.append(write_text_file(
        "insurance_claim_rodriguez_hassan.txt",
        make_insurance_claim(c[4], c[5]),
    ))
    metadata_rows.append(write_text_file(
        "insurance_claim_kim_patel.txt",
        make_insurance_claim(c[7], c[6]),
    ))
    metadata_rows.append(write_text_file(
        "payroll_register_q1_2024.txt",
        make_payroll_register([c[1], c[2], c[3], c[4], c[5]]),
    ))
    metadata_rows.append(write_text_file(
        "payroll_register_q2_2024.txt",
        make_payroll_register([c[6], c[7], c[8], c[9], c[10]]),
    ))
    metadata_rows.append(write_text_file(
        "tax_w2_hassan_2023.txt",
        make_tax_w2(c[5]),
    ))
    metadata_rows.append(write_text_file(
        "tax_w2_patel_2023.txt",
        make_tax_w2(c[6]),
    ))
    metadata_rows.append(write_text_file(
        "benefits_enrollment_batch1.txt",
        make_benefits_enrollment(c[1], c[3], c[8]),
    ))
    metadata_rows.append(write_text_file(
        "benefits_enrollment_batch2.txt",
        make_benefits_enrollment(c[4], c[7], c[10]),
    ))
    metadata_rows.append(write_text_file(
        "client_intake_whitfield.txt",
        make_client_intake(c[10]),
    ))
    metadata_rows.append(write_text_file(
        "client_intake_reyes_morales.txt",
        make_client_intake(c[9]),
    ))

    # --- .csv files (5) ---
    metadata_rows.append(write_csv_file(
        "employee_directory_all.csv",
        make_csv_employee_directory(CUSTOMERS),
    ))
    metadata_rows.append(write_csv_file(
        "payroll_export_q1.csv",
        make_csv_payroll_export(CUSTOMERS[:5]),
    ))
    metadata_rows.append(write_csv_file(
        "payroll_export_q2.csv",
        make_csv_payroll_export(CUSTOMERS[5:]),
    ))
    metadata_rows.append(write_csv_file(
        "insurance_roster_2024.csv",
        make_csv_insurance_roster(CUSTOMERS),
    ))
    metadata_rows.append(write_csv_file(
        "client_list_lastfirst.csv",
        make_csv_client_list(CUSTOMERS),
    ))

    # --- .xlsx files (5) ---
    metadata_rows.append(write_xlsx_file(
        "hr_form_smith_jones.xlsx",
        make_xlsx_hr_form(c[2]),
    ))
    metadata_rows.append(write_xlsx_file(
        "hr_form_kim.xlsx",
        make_xlsx_hr_form(c[7]),
    ))
    metadata_rows.append(write_xlsx_file(
        "payroll_q1_xlsx.xlsx",
        make_xlsx_payroll(CUSTOMERS[:5]),
    ))
    metadata_rows.append(write_xlsx_file(
        "benefits_enrollment.xlsx",
        make_xlsx_benefits(CUSTOMERS),
    ))
    metadata_rows.append(write_xlsx_file(
        "client_intake_batch.xlsx",
        make_xlsx_client_intake(CUSTOMERS),
    ))

    # --- .xls files (3) ---
    metadata_rows.append(write_xls_file(
        "payroll_legacy_q1.xls",
        make_xls_payroll(CUSTOMERS[:5]),
    ))
    metadata_rows.append(write_xls_file(
        "employee_directory_legacy.xls",
        make_xls_employee_directory(CUSTOMERS),
    ))
    metadata_rows.append(write_xls_file(
        "insurance_claims_legacy.xls",
        make_xls_insurance_claims(CUSTOMERS),
    ))

    print(f"  Breach files written: {len(metadata_rows)}")

    # ------------------------------------------------------------------
    # 4. Write dlu_metadata.csv (V2: only MD5 + file_path)
    # ------------------------------------------------------------------
    dlu_columns = ["MD5", "file_path"]
    with open(DLU_METADATA_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=dlu_columns)
        writer.writeheader()
        writer.writerows(metadata_rows)

    print(f"  dlu_metadata.csv written: {len(metadata_rows)} rows")
    print("Done.")


if __name__ == "__main__":
    generate_all()
