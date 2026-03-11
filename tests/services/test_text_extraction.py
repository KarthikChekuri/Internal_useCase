"""Tests for app.services.text_extraction — Phase 2.1."""

import csv
import io
import struct

import openpyxl
import pytest

from app.services.text_extraction import extract_text


# ── 1. .txt file ──────────────────────────────────────────────────────

class TestTxtExtraction:
    def test_extracts_utf8_text(self, tmp_path):
        f = tmp_path / "sample.txt"
        f.write_text("Hello, world!\nLine two.", encoding="utf-8")

        result = extract_text(str(f))

        assert result == "Hello, world!\nLine two."

    def test_encoding_error_returns_none(self, tmp_path):
        """Non-UTF-8 bytes that cannot be decoded should yield None."""
        f = tmp_path / "bad.txt"
        f.write_bytes(b"\x80\x81\x82\x83")

        result = extract_text(str(f))

        assert result is None

    def test_empty_txt_returns_empty_string(self, tmp_path):
        f = tmp_path / "empty.txt"
        f.write_text("", encoding="utf-8")

        result = extract_text(str(f))

        assert result == ""


# ── 2. .xlsx file ─────────────────────────────────────────────────────

class TestXlsxExtraction:
    def test_extracts_all_sheets_and_cells(self, tmp_path):
        f = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Name"
        ws1["B1"] = "Age"
        ws1["A2"] = "Alice"
        ws1["B2"] = 30

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "City"
        ws2["A2"] = "London"

        wb.save(str(f))

        result = extract_text(str(f))

        assert result is not None
        for word in ("Name", "Age", "Alice", "30", "City", "London"):
            assert word in result

    def test_empty_xlsx_returns_empty_string(self, tmp_path):
        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        # Default sheet with no data
        wb.save(str(f))

        result = extract_text(str(f))

        assert result == ""

    def test_corrupt_xlsx_returns_none(self, tmp_path):
        f = tmp_path / "corrupt.xlsx"
        f.write_bytes(b"this is not a valid xlsx file at all")

        result = extract_text(str(f))

        assert result is None


# ── 3. .xls file ─────────────────────────────────────────────────────

class TestXlsExtraction:
    def _write_minimal_xls(self, path, sheets: dict[str, list[list]]):
        """Write a real .xls file using xlwt."""
        import xlwt

        wb = xlwt.Workbook()
        for sheet_name, rows in sheets.items():
            ws = wb.add_sheet(sheet_name)
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    ws.write(r, c, val)
        wb.save(str(path))

    def test_extracts_all_sheets_and_cells(self, tmp_path):
        f = tmp_path / "data.xls"
        self._write_minimal_xls(f, {
            "Sheet1": [["Name", "Age"], ["Bob", 25]],
            "Sheet2": [["Country"], ["France"]],
        })

        result = extract_text(str(f))

        assert result is not None
        for word in ("Name", "Age", "Bob", "25", "Country", "France"):
            assert word in result

    def test_empty_xls_returns_empty_string(self, tmp_path):
        f = tmp_path / "empty.xls"
        self._write_minimal_xls(f, {"Sheet1": []})

        result = extract_text(str(f))

        assert result == ""

    def test_corrupt_xls_returns_none(self, tmp_path):
        f = tmp_path / "corrupt.xls"
        f.write_bytes(b"not a real xls file")

        result = extract_text(str(f))

        assert result is None


# ── 4. .csv file ──────────────────────────────────────────────────────

class TestCsvExtraction:
    def test_extracts_all_rows_and_columns(self, tmp_path):
        f = tmp_path / "data.csv"
        with open(str(f), "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["Name", "Email"])
            writer.writerow(["Alice", "alice@example.com"])
            writer.writerow(["Bob", "bob@example.com"])

        result = extract_text(str(f))

        assert result is not None
        for word in ("Name", "Email", "Alice", "alice@example.com", "Bob", "bob@example.com"):
            assert word in result

    def test_empty_csv_returns_empty_string(self, tmp_path):
        f = tmp_path / "empty.csv"
        f.write_text("", encoding="utf-8")

        result = extract_text(str(f))

        assert result == ""


# ── 5. Unsupported extension ─────────────────────────────────────────

class TestUnsupportedExtension:
    def test_returns_none_for_unsupported_extension(self, tmp_path):
        f = tmp_path / "image.png"
        f.write_bytes(b"\x89PNG\r\n")

        result = extract_text(str(f))

        assert result is None

    def test_returns_none_for_no_extension(self, tmp_path):
        f = tmp_path / "README"
        f.write_text("some content", encoding="utf-8")

        result = extract_text(str(f))

        assert result is None


# ── 6. File not found ────────────────────────────────────────────────

class TestFileNotFound:
    def test_returns_none_for_missing_file(self, tmp_path):
        result = extract_text(str(tmp_path / "does_not_exist.txt"))

        assert result is None


# ── 7. No exceptions raised ──────────────────────────────────────────

class TestNoExceptionsRaised:
    """extract_text must never raise — it returns None on any error."""

    def test_does_not_raise_on_missing_file(self, tmp_path):
        # Should not raise; just returns None
        result = extract_text(str(tmp_path / "nope.xlsx"))
        assert result is None

    def test_does_not_raise_on_corrupt_file(self, tmp_path):
        f = tmp_path / "bad.xlsx"
        f.write_bytes(b"\x00\x01\x02")
        result = extract_text(str(f))
        assert result is None


# ── 8. Additional edge-case tests (Phase 6.1) ────────────────────────

class TestTextExtractionEdgeCases:
    """Additional edge-case tests for text extraction robustness."""

    def test_txt_with_unicode_content(self, tmp_path):
        """UTF-8 text with unicode characters extracts correctly."""
        f = tmp_path / "unicode.txt"
        f.write_text("Jose Garcia lives at 123 Main St", encoding="utf-8")
        result = extract_text(str(f))
        assert result is not None
        assert "Jose" in result

    def test_csv_with_commas_in_quoted_fields(self, tmp_path):
        """CSV with commas inside quoted fields extracts correctly."""
        f = tmp_path / "quoted.csv"
        f.write_text(
            '"Name","Address"\n"Doe, John","123 Main St, Apt 4"\n',
            encoding="utf-8",
        )
        result = extract_text(str(f))
        assert result is not None
        assert "Doe, John" in result
        assert "123 Main St, Apt 4" in result

    def test_txt_large_file(self, tmp_path):
        """Large text file extracts without error."""
        f = tmp_path / "large.txt"
        content = "Line of text with some data.\n" * 10000
        f.write_text(content, encoding="utf-8")
        result = extract_text(str(f))
        assert result is not None
        assert len(result) > 10000

    def test_csv_single_column(self, tmp_path):
        """CSV with a single column extracts correctly."""
        f = tmp_path / "single.csv"
        f.write_text("Name\nAlice\nBob\nCharlie\n", encoding="utf-8")
        result = extract_text(str(f))
        assert result is not None
        assert "Alice" in result
        assert "Bob" in result
        assert "Charlie" in result

    def test_unsupported_pdf_extension(self, tmp_path):
        """PDF extension returns None (unsupported)."""
        f = tmp_path / "report.pdf"
        f.write_bytes(b"%PDF-1.4 fake content")
        result = extract_text(str(f))
        assert result is None

    def test_unsupported_docx_extension(self, tmp_path):
        """DOCX extension returns None (unsupported)."""
        f = tmp_path / "document.docx"
        f.write_bytes(b"fake docx content")
        result = extract_text(str(f))
        assert result is None

    def test_case_insensitive_extension(self, tmp_path):
        """Extension matching should be case-insensitive (.TXT works like .txt)."""
        f = tmp_path / "sample.TXT"
        f.write_text("Hello world", encoding="utf-8")
        result = extract_text(str(f))
        assert result == "Hello world"

    def test_xlsx_with_numeric_cells(self, tmp_path):
        """XLSX with numeric cells converts them to strings."""
        import openpyxl

        f = tmp_path / "numbers.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 12345
        ws["B1"] = 67.89
        wb.save(str(f))

        result = extract_text(str(f))
        assert result is not None
        assert "12345" in result
